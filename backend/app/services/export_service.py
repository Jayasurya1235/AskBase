"""
Generates downloadable Excel and PDF files from query results.

Takes plain data (columns + rows) — no database access here, this
is purely a formatting/export concern, kept separate from execution.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

from app.models.export import ExportRequest


def generate_excel(data: ExportRequest) -> io.BytesIO:
    """
    Builds an .xlsx file in memory (no disk write needed) and returns
    it as a byte stream ready to send as an HTTP response.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header row, styled
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(data.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_idx, row in enumerate(data.rows, start=2):
        for col_idx, col_name in enumerate(data.columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    # Auto-size columns roughly, based on content length
    for col_idx, col_name in enumerate(data.columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(row.get(col_name, ""))) for row in data.rows]
        )
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_pdf(data: ExportRequest) -> io.BytesIO:
    """
    Builds a .pdf file in memory summarizing the question and results
    as a formatted table.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("AskBase Report", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Question: {data.question}", styles["Normal"]))
    elements.append(Spacer(1, 16))

    # Build table data: header row + all data rows, as plain strings
    table_data = [data.columns]
    for row in data.rows:
        table_data.append([str(row.get(col, "")) for col in data.columns])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B5CF6")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer